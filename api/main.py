import os
from pathlib import Path

from fastapi import FastAPI, Form, HTTPException, Request
from fastapi.responses import Response, HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.sessions import SessionMiddleware

from .queries import (
    get_programa, get_productos, get_sectores,
    get_programas_semana, get_semanas_disponibles, get_productos_multiples, get_sectores_multiples,
    get_temporadas, get_sucursales, get_especies, get_variedades,
    listar_cuarteles_con_programas, agrupar_por_sucursal,
    get_cuartel_info, get_semanas_cuartel, get_productos_asignados, build_matriz,
    get_vigores, get_factores_all, get_estimaciones_cuartel,
    get_ur_cuartel, get_cuarteles_navegables, calcular_unidades, save_unidades_requeridas,
    get_resumen_costos, get_comparativa_variedad,
    get_analisis_agronomico, save_analisis_agronomico, recalcular_ur_con_ajuste,
    tiene_ajuste_agronomico, get_cuarteles_con_ajuste_temporada,
    listar_cuarteles_con_ajustes,
    get_descuento_bodega_semana,
    get_adquisiciones_consolidado, get_semanas_temporada,
    save_vigor, save_factor,
    get_programas_cuartel, get_productos_disponibles,
    agregar_producto_semanas, update_dosis, eliminar_producto_cuartel,
    get_semanas_disponibles_cuartel, agregar_semana_programa, eliminar_semana_programa,
    get_productos_lista, get_unidades_lista, save_producto, update_producto_nutrientes,
    get_objetivos, get_modos_accion,
    existe_producto_por_nombre, eliminar_producto as eliminar_producto_db,
    get_ingredientes_activos, get_actividades_producto, crear_ingrediente_activo,
    get_ias_de_producto, save_ias_de_producto, update_producto_general,
    get_papeleta_campo_rows, get_cuarteles_huerfanos, get_sucursal_info, get_semana_info,
    get_casetas_con_programa, get_caseta_info,
    validar_login, get_sucursales_permitidas, get_usuario_por_email,
    listar_usuarios_con_sucursales, actualizar_rol_usuario, set_sucursales_usuario,
    crear_usuario, eliminar_usuario, existe_usuario_nombre,
    actualizar_email_usuario,
)
from .pdf_service import build_pdf, build_pdf_bodega, build_pdf_campo

app = FastAPI(title="LH Fertilizaciones")

BASE_DIR = Path(__file__).resolve().parent
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=BASE_DIR / "templates")


# ══ AUTH MIDDLEWARE ═══════════════════════════════════════════════════════════

PUBLIC_PATHS = {"/login", "/logout", "/health", "/", "/docs", "/openapi.json", "/redoc", "/dev-login"}
PUBLIC_PREFIXES = ("/static", "/papeleta", "/registro-semanal", "/login/")


class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        is_public = (
            path in PUBLIC_PATHS
            or any(path.startswith(p) for p in PUBLIC_PREFIXES)
        )
        if is_public or request.session.get("user_id"):
            return await call_next(request)
        return RedirectResponse(url=f"/login?next={path}", status_code=303)


class ContextMiddleware(BaseHTTPMiddleware):
    """Inyecta sucursal activa + listado de sucursales en request.state para los templates."""
    async def dispatch(self, request: Request, call_next):
        suc = request.session.get("id_sucursal")
        request.state.id_sucursal_activa = int(suc) if suc else None
        if not hasattr(app.state, "sucursales_cache") or not app.state.sucursales_cache:
            try:
                app.state.sucursales_cache = get_sucursales()
            except Exception:
                app.state.sucursales_cache = []
        # Filtrar segun permisos: None = sin restriccion (admin / super_admin).
        # Defensive: si un admin viene de una sesion vieja con [] en lugar de
        # None, NO filtrar a nada.
        permitidas = request.session.get("user_sucursales")
        rol = request.session.get("user_rol") or "user"
        if permitidas is None or rol in ("admin", "super_admin"):
            request.state.sucursales_all = app.state.sucursales_cache
        else:
            permitidas_set = set(permitidas)
            request.state.sucursales_all = [
                s for s in app.state.sucursales_cache if s["id"] in permitidas_set
            ]
        return await call_next(request)


# Orden de add_middleware: el primero agregado es el mas interior.
# Ejecucion entrante: Session -> Auth -> Context -> endpoint
app.add_middleware(ContextMiddleware)
app.add_middleware(AuthMiddleware)
app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv("SESSION_SECRET_KEY", "dev-insecure-change-me"),
    session_cookie="ferti_session",
    max_age=60 * 60 * 12,  # 12h
    same_site="lax",
    # En produccion (Cloud Run) la app esta detras de HTTPS. Marcar la cookie
    # como Secure evita problemas con flujos OAuth donde el state se valida
    # entre el authorize_redirect y el callback. En local con http, dejar False.
    https_only=os.getenv("COOKIE_SECURE", "1") == "1",
)


def _id_sucursal(request: Request, query_value: int | None) -> int | None:
    """Prioridad: query param explicito > sucursal en sesion > None.

    Si el usuario tiene restriccion (user no-admin con sucursales asignadas),
    valida que el id resultante este permitido; si no lo esta cae a la primera
    permitida. Si no eligio ninguna y tiene permisos, fuerza la primera.
    """
    if query_value is not None:
        candidato = query_value
    else:
        s = request.session.get("id_sucursal")
        candidato = int(s) if s else None

    permitidas = request.session.get("user_sucursales")
    if permitidas is None:
        return candidato  # admin, sin restriccion
    permitidas_set = set(permitidas)
    if not permitidas_set:
        return candidato  # user sin sucursales asignadas: no restringir (no romper)
    if candidato is not None and candidato in permitidas_set:
        return candidato
    # candidato fuera del set o sin candidato -> primera permitida
    return next(iter(sorted(permitidas_set)))


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/", include_in_schema=False)
def root(request: Request):
    if request.session.get("user_id"):
        return RedirectResponse(url="/app/programas", status_code=303)
    return RedirectResponse(url="/login", status_code=303)


# ══ LOGIN / LOGOUT ════════════════════════════════════════════════════════════

# ══ GOOGLE OAUTH ══════════════════════════════════════════════════════════════
# Workspace exclusivo: hd=lahornilla.cl evita que cualquier Gmail entre al
# consentimiento. Igual validamos el dominio del email en el callback porque
# 'hd' es un hint de UI, no una garantia.
WORKSPACE_DOMAIN = "lahornilla.cl"

from authlib.integrations.starlette_client import OAuth, OAuthError  # noqa: E402

oauth = OAuth()
oauth.register(
    name="google",
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_id=os.getenv("GOOGLE_CLIENT_ID"),
    client_secret=os.getenv("GOOGLE_CLIENT_SECRET"),
    client_kwargs={
        "scope": "openid email profile",
        "prompt": "select_account",
    },
)


def _login_session_for(request: Request, user: dict, next_path: str) -> RedirectResponse:
    """Setea la sesion despues de validar al usuario (login local u OAuth)."""
    nombre = user.get("nombre") or ""
    apellido = user.get("apellido") or ""
    request.session["user_id"] = user["id"]
    request.session["user_usuario"] = user["usuario"]
    request.session["user_name"] = (nombre + " " + apellido).strip() or user["usuario"]
    request.session["user_initials"] = (
        (nombre[:1] + apellido[:1]).upper() or user["usuario"][:2].upper()
    )
    user_rol = user.get("rol") or "user"
    request.session["user_rol"] = user_rol
    if user_rol in ("admin", "super_admin"):
        request.session["user_sucursales"] = None
    else:
        permitidas = get_sucursales_permitidas(user["id"])
        request.session["user_sucursales"] = permitidas
        if permitidas:
            request.session["id_sucursal"] = sorted(permitidas)[0]
    destino = next_path if next_path and next_path.startswith("/") else "/app/programas"
    return RedirectResponse(url=destino, status_code=303)


@app.get("/login", response_class=HTMLResponse)
def web_login(request: Request, next: str = "/app/programas", error: str | None = None):
    if request.session.get("user_id"):
        return RedirectResponse(url=next or "/app/programas", status_code=303)
    errores = {
        "dominio": "Tu cuenta no pertenece al dominio @lahornilla.cl.",
        "no_registrado": "Tu correo no esta registrado en el sistema. Contacta a TI.",
        "oauth": "No se pudo completar el inicio de sesion. Intenta de nuevo.",
    }
    return templates.TemplateResponse(
        "login.html",
        {"request": request, "next": next, "error": errores.get(error or "")},
    )


@app.get("/login/google")
async def login_google(request: Request, next: str = "/app/programas"):
    # Guardar destino post-login en la sesion (el state de OAuth no admite payload propio aca).
    request.session["oauth_next"] = next if next and next.startswith("/") else "/app/programas"
    redirect_uri = str(request.url_for("auth_google_callback"))
    return await oauth.google.authorize_redirect(request, redirect_uri, hd=WORKSPACE_DOMAIN)


@app.get("/login/google/callback", name="auth_google_callback")
async def auth_google_callback(request: Request):
    try:
        token = await oauth.google.authorize_access_token(request)
    except OAuthError as e:
        import logging
        logging.error("OAuth callback fallo: %s | description=%s | session_keys=%s",
                      e.error, e.description, list(request.session.keys()))
        return RedirectResponse(url="/login?error=oauth", status_code=303)
    except Exception as e:
        import logging, traceback
        logging.error("OAuth callback excepcion: %s\n%s", e, traceback.format_exc())
        return RedirectResponse(url="/login?error=oauth", status_code=303)
    userinfo = token.get("userinfo") or {}
    email = (userinfo.get("email") or "").lower()
    if not email or not userinfo.get("email_verified"):
        return RedirectResponse(url="/login?error=oauth", status_code=303)
    # Validar dominio Workspace (defensa: hd es solo un hint en la UI de Google)
    if not email.endswith(f"@{WORKSPACE_DOMAIN}"):
        return RedirectResponse(url="/login?error=dominio", status_code=303)
    user = get_usuario_por_email(email)
    if not user:
        return RedirectResponse(url="/login?error=no_registrado", status_code=303)
    next_path = request.session.pop("oauth_next", "/app/programas")
    return _login_session_for(request, user, next_path)


@app.get("/logout")
def do_logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


@app.get("/dev-login")
def dev_login(request: Request, email: str = "fsoto@lahornilla.cl", next: str = "/app/programas"):
    """Bypass de OAuth solo para desarrollo local. Requiere DEV_MODE=1 en el
    entorno. En produccion siempre devuelve 404."""
    if os.getenv("DEV_MODE") != "1":
        raise HTTPException(status_code=404, detail="Not found")
    user = get_usuario_por_email(email.lower())
    if not user:
        raise HTTPException(status_code=404, detail=f"Usuario {email} no existe")
    return _login_session_for(request, user, next)




@app.post("/set-sucursal")
def set_sucursal(
    request: Request,
    id_sucursal: str = Form(""),
    next: str = Form("/app/programas"),
):
    if not id_sucursal or id_sucursal in ("none", "0"):
        request.session.pop("id_sucursal", None)
    else:
        try:
            id_int = int(id_sucursal)
            permitidas = request.session.get("user_sucursales")
            # Si tiene restriccion y no esta en el set, ignorar
            if permitidas is None or id_int in set(permitidas):
                request.session["id_sucursal"] = id_int
        except ValueError:
            pass
    destino = next if next and next.startswith("/") else "/app/programas"
    return RedirectResponse(url=destino, status_code=303)


def _id_responsable(request: Request) -> int:
    return int(request.session.get("user_id") or 0)


def _es_admin(request: Request) -> bool:
    return (request.session.get("user_rol") or "user") in ("admin", "super_admin")


def _es_super_admin(request: Request) -> bool:
    return (request.session.get("user_rol") or "user") == "super_admin"


def _require_admin(request: Request) -> None:
    if not _es_admin(request):
        raise HTTPException(status_code=403, detail="Acceso restringido a administradores.")


def _require_super_admin(request: Request) -> None:
    if not _es_super_admin(request):
        raise HTTPException(status_code=403, detail="Acceso restringido a super administradores.")


def _require_cuartel_permitido(request: Request, id_cuartel: int) -> None:
    """Tira 403 si el cuartel pertenece a una sucursal fuera de los permisos del usuario.
    Admin y super_admin no tienen restriccion. Cache por request para evitar queries
    repetidas en endpoints que se invocan varias veces."""
    permitidas = request.session.get("user_sucursales")
    if permitidas is None:
        return  # admin / super_admin
    cache = getattr(request.state, "_cuartel_suc_cache", None)
    if cache is None:
        cache = {}
        request.state._cuartel_suc_cache = cache
    if id_cuartel in cache:
        id_suc = cache[id_cuartel]
    else:
        from .queries import get_sucursal_de_cuartel
        id_suc = get_sucursal_de_cuartel(id_cuartel)
        cache[id_cuartel] = id_suc
    if id_suc is None or id_suc not in set(permitidas):
        raise HTTPException(status_code=403, detail="Cuartel no autorizado")


# ══ WEB APP ═══════════════════════════════════════════════════════════════════

@app.get("/app", include_in_schema=False)
def app_root():
    return RedirectResponse(url="/app/programas")


def _to_int(v):
    """Convierte string vacio o None a None, sino a int."""
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


@app.get("/app/programas", response_class=HTMLResponse)
def web_programas(
    request: Request,
    temporada: str | None = None,
    sucursal: str | None = None,
    especie: str | None = None,
    variedad: str | None = None,
):
    # Tolerar query params vacios (?variedad= que mande el form al limpiar)
    temporada_id = _to_int(temporada)
    sucursal_id = _to_int(sucursal)
    especie_id = _to_int(especie)
    variedad_id = _to_int(variedad)

    temporadas = get_temporadas()
    id_suc = _id_sucursal(request, sucursal_id)

    # Especies y variedades acotadas a la sucursal activa
    especies_disponibles = get_especies(id_sucursal=id_suc)
    variedades_de_especie = get_variedades(id_especie=especie_id, id_sucursal=id_suc)

    # Si la variedad seleccionada no esta en el filtro actual, descartarla
    if variedad_id:
        ids_validos = {v["id"] for v in variedades_de_especie}
        if variedad_id not in ids_validos:
            variedad_id = None

    # Si la especie seleccionada no esta en las disponibles, descartarla
    if especie_id:
        ids_esp_validos = {e["id"] for e in especies_disponibles}
        if especie_id not in ids_esp_validos:
            especie_id = None
            variedad_id = None
            variedades_de_especie = get_variedades(id_sucursal=id_suc)

    cuarteles = listar_cuarteles_con_programas(
        id_temporada=temporada_id,
        id_sucursal=id_suc,
        id_especie=especie_id,
        id_variedad=variedad_id,
    )

    # Marcar cuarteles con ajuste agronomico activo para mostrar icono
    id_temp_para_ajustes = temporada_id or (temporadas[0]["id"] if temporadas else None)
    cuarteles_con_ajuste = (
        get_cuarteles_con_ajuste_temporada(id_temp_para_ajustes)
        if id_temp_para_ajustes else set()
    )
    for c in cuarteles:
        c["tiene_ajuste"] = c["id_cuartel"] in cuarteles_con_ajuste

    grupos = agrupar_por_sucursal(cuarteles)
    semanas = get_semanas_disponibles(id_temporada=temporada_id, id_sucursal=id_suc)
    return templates.TemplateResponse(
        "programas.html",
        {
            "request": request,
            "active_page": "programas",
            "temporadas": temporadas,
            "sucursales": request.state.sucursales_all,
            "especies": especies_disponibles,
            "variedades": variedades_de_especie,
            "grupos": grupos,
            "programas": cuarteles,
            "total_cuarteles": len(cuarteles),
            "filtro_temporada": temporada_id,
            "filtro_sucursal": id_suc,
            "filtro_especie": especie_id,
            "filtro_variedad": variedad_id,
            "semanas": semanas,
        },
    )


@app.get("/app/matriz", response_class=HTMLResponse)
def web_matriz_root():
    return RedirectResponse(url="/app/programas")


@app.get("/app/matriz/{id_cuartel}", response_class=HTMLResponse)
def web_matriz(
    request: Request,
    id_cuartel: int,
    temporada: int | None = None,
    err: str | None = None,
):
    _require_cuartel_permitido(request, id_cuartel)
    cuartel = get_cuartel_info(id_cuartel)
    if not cuartel:
        raise HTTPException(status_code=404, detail="Cuartel no encontrado")

    temporadas = get_temporadas()
    id_temp = temporada or (temporadas[0]["id"] if temporadas else None)

    semanas_rows = get_semanas_cuartel(id_cuartel, id_temporada=temporada)
    ids_prog = [s["id_programa"] for s in semanas_rows]
    productos_rows = get_productos_asignados(ids_prog)
    matriz = build_matriz(semanas_rows, productos_rows)
    ur = get_ur_cuartel(id_cuartel, id_temp)
    analisis = get_analisis_agronomico(id_cuartel, id_temp) if id_temp else None

    # Cuarteles navegables para el selector: prioriza la sucursal seleccionada
    # en la sesion (para que el listado se actualice cuando el usuario cambia
    # de sucursal en el topbar). Si no hay ninguna en sesion, cae a la del
    # cuartel actual.
    id_suc_nav = _id_sucursal(request, None) or cuartel.get("id_sucursal")
    cuarteles_nav = get_cuarteles_navegables(id_suc_nav, id_temp)

    errores = {
        "semana_duplicada": "Esa semana ya estaba en el programa. No se duplicó.",
    }
    alert_msg = errores.get(err) if err else None

    return templates.TemplateResponse(
        "matriz.html",
        {
            "request": request,
            "active_page": "matriz",
            "cuartel": cuartel,
            "matriz": matriz,
            "ur": ur,
            "analisis": analisis,
            "temporadas": temporadas,
            "filtro_temporada": temporada,
            "cuarteles_nav": cuarteles_nav,
            "alert_msg": alert_msg,
        },
    )


# ── Unidades Requeridas ──────────────────────────────────────────────────────

@app.get("/app/unidades-requeridas", response_class=HTMLResponse)
def web_listado_ur(
    request: Request,
    temporada: int | None = None,
    sucursal: int | None = None,
    estado: str = "sin_ur",
):
    if estado not in ("con_ur", "sin_ur", "todos"):
        estado = "sin_ur"
    id_suc = _id_sucursal(request, sucursal)
    cuarteles = listar_cuarteles_con_programas(
        id_temporada=temporada, id_sucursal=id_suc, filtro_ur=estado
    )
    return templates.TemplateResponse(
        "unidades_listado.html",
        {
            "request": request,
            "active_page": "unidades-req",
            "temporadas": get_temporadas(),
            "sucursales": request.state.sucursales_all,
            "grupos": agrupar_por_sucursal(cuarteles),
            "total_cuarteles": len(cuarteles),
            "filtro_temporada": temporada,
            "filtro_sucursal": id_suc,
            "filtro_estado": estado,
        },
    )


@app.get("/app/unidades/{id_cuartel}", response_class=HTMLResponse)
def web_unidades(request: Request, id_cuartel: int, temporada: int | None = None):
    _require_cuartel_permitido(request, id_cuartel)
    cuartel = get_cuartel_info(id_cuartel)
    if not cuartel:
        raise HTTPException(status_code=404, detail="Cuartel no encontrado")

    temporadas = get_temporadas()
    estimaciones = get_estimaciones_cuartel(id_cuartel)

    # UR guardada actual (para pre-seleccionar vigor)
    id_temp_actual = temporada or (temporadas[0]["id"] if temporadas else None)
    ur_actual = get_ur_cuartel(id_cuartel, id_temp_actual) if id_temp_actual else None

    return templates.TemplateResponse(
        "unidades.html",
        {
            "request": request,
            "active_page": "programas",
            "cuartel": cuartel,
            "temporadas": temporadas,
            "vigores": get_vigores(),
            "estimaciones": estimaciones,
            "filtro_temporada": temporada,
            "ur_actual": ur_actual,
        },
    )


def _aplicar_ajuste_y_regenerar_ur(
    request: Request,
    id_cuartel: int,
    id_temporada: int,
    factores: dict,
) -> None:
    """Upsert del ajuste y, si ya hay UR para ese cuartel/temporada, la regenera."""
    save_analisis_agronomico(
        id_cuartel=id_cuartel,
        id_temporada=id_temporada,
        factores=factores,
        id_responsable=_id_responsable(request),
    )
    ur = get_ur_cuartel(id_cuartel, id_temporada)
    if ur:
        estimaciones = get_estimaciones_cuartel(id_cuartel)
        if estimaciones:
            est = estimaciones[0]
            vigores = get_vigores()
            vigor = next((v for v in vigores if v["id"] == ur.get("id_vigor")), None)
            if vigor:
                save_unidades_requeridas(
                    id_cuartel=id_cuartel,
                    id_temporada=id_temporada,
                    id_vigor=ur["id_vigor"],
                    id_responsable=_id_responsable(request),
                    especie=est["especie"],
                    ton_estimadas=float(est["ton_estimadas"]),
                    vigor_factor=float(vigor["factor"]),
                    factores=get_factores_all(),
                )


@app.get("/app/ajuste-agronomico", response_class=HTMLResponse)
def web_ajuste_agronomico(
    request: Request,
    temporada: str | None = None,
    sucursal: str | None = None,
    especie: str | None = None,
    variedad: str | None = None,
    cuartel: int | None = None,
):
    # Acceso para todos los usuarios autenticados. Los users restringidos
    # solo veran cuarteles de las sucursales que tienen asignadas (via
    # _id_sucursal + listar_cuarteles_con_ajustes con id_sucursal).
    temporada_id = _to_int(temporada)
    sucursal_id = _to_int(sucursal)
    especie_id = _to_int(especie)
    variedad_id = _to_int(variedad)

    temporadas = get_temporadas()
    id_temp = temporada_id or (temporadas[0]["id"] if temporadas else None)
    id_suc = _id_sucursal(request, sucursal_id)

    especies_disponibles = get_especies(id_sucursal=id_suc)
    variedades_de_especie = get_variedades(id_especie=especie_id, id_sucursal=id_suc)

    if id_temp:
        filas = listar_cuarteles_con_ajustes(
            id_temporada=id_temp,
            id_sucursal=id_suc,
            id_especie=especie_id,
            id_variedad=variedad_id,
        )
    else:
        filas = []

    return templates.TemplateResponse(
        "ajuste_agronomico.html",
        {
            "request": request,
            "active_page": "ajuste-agronomico",
            "temporadas": temporadas,
            "especies": especies_disponibles,
            "variedades": variedades_de_especie,
            "filas": filas,
            "filtro_temporada": id_temp,
            "filtro_sucursal": id_suc,
            "filtro_especie": especie_id,
            "filtro_variedad": variedad_id,
            "cuartel_destacado": cuartel,
        },
    )


@app.post("/app/ajuste-agronomico/{id_cuartel}")
def post_ajuste_agronomico_fila(
    request: Request,
    id_cuartel: int,
    temporada: int = Form(...),
    factor_n: float = Form(1.0),
    factor_p: float = Form(1.0),
    factor_k: float = Form(1.0),
    factor_mg: float = Form(1.0),
    factor_b: float = Form(1.0),
    factor_ca: float = Form(1.0),
    factor_zn: float = Form(1.0),
    factor_mn: float = Form(1.0),
):
    """Autosave por fila desde la tabla de ajuste nutricional. Responde 204."""
    _require_cuartel_permitido(request, id_cuartel)
    _aplicar_ajuste_y_regenerar_ur(
        request,
        id_cuartel,
        temporada,
        {
            "n": factor_n, "k": factor_k, "p": factor_p, "mg": factor_mg,
            "b": factor_b, "ca": factor_ca, "zn": factor_zn, "mn": factor_mn,
        },
    )
    return HTMLResponse(status_code=204)


@app.get("/app/unidades/{id_cuartel}/preview", response_class=HTMLResponse)
def preview_unidades(
    request: Request,
    id_cuartel: int,
    id_estimacion: str,
    id_vigor: int,
):
    _require_cuartel_permitido(request, id_cuartel)
    estimaciones = get_estimaciones_cuartel(id_cuartel)
    est = next((e for e in estimaciones if str(e["id_estimacion"]) == id_estimacion), None)
    if not est:
        return HTMLResponse("")

    vigores = get_vigores()
    vigor = next((v for v in vigores if v["id"] == id_vigor), None)
    if not vigor:
        return HTMLResponse("")

    factores = get_factores_all()
    unidades = calcular_unidades(
        float(est["ton_estimadas"]),
        float(vigor["factor"]),
        est["especie"],
        factores,
    )

    return templates.TemplateResponse(
        "unidades_preview.html",
        {"request": request, "unidades": unidades, "estimacion": est, "vigor": vigor},
    )


@app.post("/app/unidades/{id_cuartel}")
def crear_unidades(
    request: Request,
    id_cuartel: int,
    id_estimacion: str = Form(...),
    id_vigor: int = Form(...),
    id_temporada: int = Form(...),
):
    _require_cuartel_permitido(request, id_cuartel)
    estimaciones = get_estimaciones_cuartel(id_cuartel)
    est = next((e for e in estimaciones if str(e["id_estimacion"]) == id_estimacion), None)
    if not est:
        raise HTTPException(status_code=400, detail="Estimación no encontrada")

    vigores = get_vigores()
    vigor = next((v for v in vigores if v["id"] == id_vigor), None)
    if not vigor:
        raise HTTPException(status_code=400, detail="Vigor no encontrado")

    factores = get_factores_all()
    save_unidades_requeridas(
        id_cuartel=id_cuartel,
        id_temporada=id_temporada,
        id_vigor=id_vigor,
        id_responsable=_id_responsable(request),
        especie=est["especie"],
        ton_estimadas=float(est["ton_estimadas"]),
        vigor_factor=float(vigor["factor"]),
        factores=factores,
    )
    return RedirectResponse(url=f"/app/matriz/{id_cuartel}?temporada={id_temporada}", status_code=303)


# ── Edición de matriz ────────────────────────────────────────────────────────

@app.get("/app/matriz/{id_cuartel}/productos-disponibles", response_class=HTMLResponse)
def productos_disponibles_fragment(request: Request, id_cuartel: int, temporada: int | None = None):
    _require_cuartel_permitido(request, id_cuartel)
    productos = get_productos_disponibles(id_cuartel, temporada)
    return templates.TemplateResponse(
        "fragment_productos_select.html",
        {"request": request, "productos": productos, "id_cuartel": id_cuartel, "temporada": temporada},
    )


@app.post("/app/matriz/{id_cuartel}/agregar-producto")
def agregar_producto(
    request: Request,
    id_cuartel: int,
    id_producto: list[str] = Form(...),
    temporada: int | None = Form(None),
):
    _require_cuartel_permitido(request, id_cuartel)
    # id_producto en BD es varchar(25) (UUID), no int
    ids_prog = get_programas_cuartel(id_cuartel, temporada)
    for pid in id_producto:
        agregar_producto_semanas(ids_prog, pid)
    url = f"/app/matriz/{id_cuartel}" + (f"?temporada={temporada}" if temporada else "")
    return RedirectResponse(url=url, status_code=303)


@app.post("/app/matriz/{id_cuartel}/eliminar-producto")
def eliminar_producto(
    request: Request,
    id_cuartel: int,
    id_producto: str = Form(...),
    temporada: int | None = Form(None),
):
    _require_cuartel_permitido(request, id_cuartel)
    ids_prog = get_programas_cuartel(id_cuartel, temporada)
    eliminar_producto_cuartel(ids_prog, id_producto)
    url = f"/app/matriz/{id_cuartel}" + (f"?temporada={temporada}" if temporada else "")
    return RedirectResponse(url=url, status_code=303)


@app.post("/app/matriz/{id_cuartel}/dosis", response_class=HTMLResponse)
def guardar_dosis(
    request: Request,
    id_cuartel: int,
    id_programa: str = Form(...),
    id_producto: str = Form(...),
    dosis: float = Form(...),
):
    _require_cuartel_permitido(request, id_cuartel)
    update_dosis(id_programa, id_producto, dosis)
    return HTMLResponse(f'<span class="celda-saved">{dosis:.0f}</span>', status_code=200)


@app.get("/app/matriz/{id_cuartel}/semanas-disponibles", response_class=HTMLResponse)
def semanas_disponibles_fragment(
    request: Request,
    id_cuartel: int,
    temporada: int | None = None,
):
    _require_cuartel_permitido(request, id_cuartel)
    temporadas = get_temporadas()
    id_temp = temporada or (temporadas[0]["id"] if temporadas else None)
    semanas = get_semanas_disponibles_cuartel(id_cuartel, id_temp) if id_temp else []
    return templates.TemplateResponse(
        "fragment_semanas_select.html",
        {
            "request": request,
            "id_cuartel": id_cuartel,
            "id_temporada": id_temp,
            "semanas": semanas,
        },
    )


@app.post("/app/matriz/{id_cuartel}/agregar-semana")
def agregar_semana(
    request: Request,
    id_cuartel: int,
    id_semana: int = Form(...),
    etapa: str = Form("PRECOSECHA"),
    temporada: int = Form(...),
):
    _require_cuartel_permitido(request, id_cuartel)
    _, created = agregar_semana_programa(
        id_cuartel=id_cuartel,
        id_temporada=temporada,
        id_semana=id_semana,
        etapa=etapa,
        id_responsable=_id_responsable(request),
    )
    err_qs = "" if created else "&err=semana_duplicada"
    return RedirectResponse(
        url=f"/app/matriz/{id_cuartel}?temporada={temporada}{err_qs}",
        status_code=303,
    )


@app.post("/app/matriz/{id_cuartel}/eliminar-semana")
def eliminar_semana(
    request: Request,
    id_cuartel: int,
    id_programa: str = Form(...),
    temporada: int | None = Form(None),
):
    _require_cuartel_permitido(request, id_cuartel)
    eliminar_semana_programa(id_programa)
    url = f"/app/matriz/{id_cuartel}" + (f"?temporada={temporada}" if temporada else "")
    return RedirectResponse(url=url, status_code=303)


@app.get("/app/matriz/{id_cuartel}/resumen-semanal/excel")
def resumen_semanal_excel(
    request: Request,
    id_cuartel: int,
    temporada: int | None = None,
):
    _require_cuartel_permitido(request, id_cuartel)
    cuartel = get_cuartel_info(id_cuartel)
    if not cuartel:
        raise HTTPException(404, "Cuartel no encontrado")
    semanas_rows = get_semanas_cuartel(id_cuartel, temporada)
    ids_prog = [s["id_programa"] for s in semanas_rows]
    productos_rows = get_productos_asignados(ids_prog)
    matriz = build_matriz(semanas_rows, productos_rows)
    ur = get_ur_cuartel(id_cuartel, temporada) if temporada else None

    NUTRIENTES = ["N", "K", "P", "Mg", "B", "Ca", "Zn", "Mn"]
    nuts_vistos = [n for n in NUTRIENTES if (matriz["totales_aporte"].get(n) or 0) > 0]
    sup = float(cuartel.get("sup_productiva") or 0)

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HDR_FILL = PatternFill("solid", fgColor="2d5a1f")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    ZEBRA = PatternFill("solid", fgColor="F4FAF1")
    TOTAL_FILL = PatternFill("solid", fgColor="D7E8C9")
    EMPTY_FILL = PatternFill("solid", fgColor="F5F5F5")
    THIN = Side(style="thin", color="C5D9BB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen semanal"

    # Encabezado del cuartel
    ws["A1"] = f"Resumen semanal — {cuartel.get('nombre') or cuartel.get('descripcion_ceco') or id_cuartel}"
    ws["A1"].font = Font(bold=True, size=12, color="2d5a1f")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    meta_parts = []
    if cuartel.get("sucursal"): meta_parts.append(str(cuartel["sucursal"]))
    if cuartel.get("variedad"): meta_parts.append(str(cuartel["variedad"]))
    if sup > 0: meta_parts.append(f"{sup:.2f} ha")
    if matriz.get("costo_ha_total"): meta_parts.append(f"${matriz['costo_ha_total']:.2f} USD/ha")
    ws["A2"] = " · ".join(meta_parts)
    ws["A2"].font = Font(italic=True, color="666666", size=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # Headers tabla (fila 4). Unidad neutra porque puede haber kg + L mezclados.
    unidades_str = " + ".join(sorted((matriz.get("totales_por_unidad") or {}).keys())) or "kg"
    headers = ["Semana", "Etapa", f"Dosis/ha ({unidades_str})"]
    if sup > 0: headers.append(f"Cantidad sup. ({unidades_str})")
    headers.append("USD/ha")
    headers.extend([f"{n} (un)" for n in nuts_vistos])

    HEADER_ROW = 4
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Filas
    row = HEADER_ROW + 1
    for fila in matriz["filas"]:
        sem = fila["semana"]
        vacia = (fila.get("total_kg") or 0) <= 0
        vals = [sem.get("etiqueta") or "", sem.get("etapa") or ""]
        vals.append(round(fila.get("total_kg") or 0, 1) if not vacia else None)
        if sup > 0:
            vals.append(round((fila.get("total_kg") or 0) * sup, 1) if not vacia else None)
        vals.append(round(fila.get("usd_ha") or 0, 2) if (fila.get("usd_ha") or 0) > 0 else None)
        for n in nuts_vistos:
            v = (fila.get("aporte") or {}).get(n) or 0
            vals.append(round(v, 2) if v > 0.01 else None)
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.border = BORDER
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="right")
            if vacia:
                cell.fill = EMPTY_FILL
            elif row % 2 == 0:
                cell.fill = ZEBRA
        row += 1

    # Total
    total_row = row
    total_vals = ["Total", ""]
    total_kg = sum((f.get("total_kg") or 0) for f in matriz["filas"])
    total_vals.append(round(total_kg, 1))
    if sup > 0: total_vals.append(round(total_kg * sup, 1))
    total_vals.append(round(matriz.get("costo_ha_total") or 0, 2))
    for n in nuts_vistos:
        total_vals.append(round(matriz["totales_aporte"].get(n) or 0, 2))
    for col_idx, v in enumerate(total_vals, start=1):
        cell = ws.cell(row=total_row, column=col_idx, value=v)
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)
        cell.border = BORDER
        if col_idx >= 3:
            cell.alignment = Alignment(horizontal="right")

    # Fila req UR si existe
    if ur and nuts_vistos:
        req_row = total_row + 1
        ws.cell(row=req_row, column=1, value="Requerimiento UR").font = Font(italic=True, color="666666", size=9)
        offset = 3 + (1 if sup > 0 else 0) + 1  # primera col de nutrientes
        for i, n in enumerate(nuts_vistos):
            req = ur.get(f"unidades_{n}") or 0
            if req:
                cell = ws.cell(row=req_row, column=offset + i, value=round(float(req), 1))
                cell.font = Font(italic=True, color="666666", size=9)
                cell.alignment = Alignment(horizontal="right")
                cell.border = BORDER

    # Anchos columnas
    widths = [14, 14, 10] + ([10] if sup > 0 else []) + [10] + [10] * len(nuts_vistos)
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=HEADER_ROW, column=col_idx).column_letter].width = w
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1).coordinate

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre_cuartel = (cuartel.get("nombre") or cuartel.get("descripcion_ceco") or f"cuartel_{id_cuartel}")
    safe = "".join(c if c.isalnum() else "_" for c in str(nombre_cuartel))[:40]
    fn = f"resumen_semanal_{safe}.xlsx"
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@app.get("/app/matriz/{id_cuartel}/programa-productos/excel")
def programa_productos_excel(
    request: Request,
    id_cuartel: int,
    temporada: int | None = None,
):
    """Excel matriz semana x producto (nombre comercial). Cada celda = cantidad
    total a aplicar esa semana (dosis/ha x sup_productiva). Formato solicitado
    por encargados de campo para armar retiros/aplicaciones."""
    _require_cuartel_permitido(request, id_cuartel)
    cuartel = get_cuartel_info(id_cuartel)
    if not cuartel:
        raise HTTPException(404, "Cuartel no encontrado")
    semanas_rows = get_semanas_cuartel(id_cuartel, temporada)
    ids_prog = [s["id_programa"] for s in semanas_rows]
    productos_rows = get_productos_asignados(ids_prog)
    matriz = build_matriz(semanas_rows, productos_rows)

    sup = float(cuartel.get("sup_productiva") or 0)
    productos = matriz.get("productos") or []

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HDR_FILL = PatternFill("solid", fgColor="2d5a1f")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    ZEBRA = PatternFill("solid", fgColor="F4FAF1")
    TOTAL_FILL = PatternFill("solid", fgColor="D7E8C9")
    EMPTY_FILL = PatternFill("solid", fgColor="F5F5F5")
    THIN = Side(style="thin", color="C5D9BB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = Workbook()
    ws = wb.active
    ws.title = "Programa por producto"

    # Encabezado
    ws["A1"] = f"Programa por producto — {cuartel.get('nombre') or cuartel.get('descripcion_ceco') or id_cuartel}"
    ws["A1"].font = Font(bold=True, size=12, color="2d5a1f")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(5, len(productos) + 1))
    meta_parts = []
    if cuartel.get("sucursal"): meta_parts.append(str(cuartel["sucursal"]))
    if cuartel.get("variedad"): meta_parts.append(str(cuartel["variedad"]))
    if sup > 0: meta_parts.append(f"{sup:.2f} ha")
    meta_parts.append("Cantidades = dosis/ha x superficie" if sup > 0 else "Cantidades = dosis/ha (sin sup cargada)")
    ws["A2"] = " · ".join(meta_parts)
    ws["A2"].font = Font(italic=True, color="666666", size=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(5, len(productos) + 1))

    if not productos:
        ws["A4"] = "Este cuartel no tiene productos asignados en el programa."
        ws["A4"].font = Font(italic=True, color="999999")
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
    else:
        # Headers: Semana + Fecha + productos + Total
        HEADER_ROW = 4
        headers = ["Semana", "Fecha"] + [f"{p['nombre']} ({p['unidad'] or 'kg'})" for p in productos] + ["Total"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

        # Multiplicador: si sup=0, cantidades = dosis/ha; si sup>0, cantidades = dosis/ha * sup
        mult = sup if sup > 0 else 1.0

        # Filas: una por semana
        totales_por_producto = [0.0] * len(productos)
        row = HEADER_ROW + 1
        for fila in matriz["filas"]:
            sem = fila["semana"]
            fecha = ""
            if sem.get("fecha_inicio") and sem.get("fecha_fin"):
                fi = sem["fecha_inicio"]
                ff = sem["fecha_fin"]
                fecha = f"{fi.strftime('%d/%m')} - {ff.strftime('%d/%m')}"
            vals = [sem.get("etiqueta") or "", fecha]
            total_sem = 0.0
            vacia = True
            for i, prod in enumerate(productos):
                dosis = fila["celdas"][i]
                if dosis and dosis > 0:
                    cant = dosis * mult
                    vals.append(round(cant, 2))
                    totales_por_producto[i] += cant
                    total_sem += cant
                    vacia = False
                else:
                    vals.append(None)
            vals.append(round(total_sem, 2) if total_sem > 0 else None)
            for col_idx, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=v)
                cell.border = BORDER
                if col_idx >= 3:
                    cell.alignment = Alignment(horizontal="right")
                if vacia:
                    cell.fill = EMPTY_FILL
                elif row % 2 == 0:
                    cell.fill = ZEBRA
            row += 1

        # Fila total
        total_row = row
        etiqueta_total = "Total superficie" if sup > 0 else "Total dosis/ha"
        total_vals = [etiqueta_total, ""]
        gran_total = 0.0
        for t in totales_por_producto:
            total_vals.append(round(t, 2) if t > 0 else None)
            gran_total += t
        total_vals.append(round(gran_total, 2) if gran_total > 0 else None)
        for col_idx, v in enumerate(total_vals, start=1):
            cell = ws.cell(row=total_row, column=col_idx, value=v)
            cell.fill = TOTAL_FILL
            cell.font = Font(bold=True)
            cell.border = BORDER
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="right")

        # Anchos
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        for i in range(len(productos)):
            col_letter = ws.cell(row=HEADER_ROW, column=3 + i).column_letter
            ws.column_dimensions[col_letter].width = 16
        total_col_letter = ws.cell(row=HEADER_ROW, column=3 + len(productos)).column_letter
        ws.column_dimensions[total_col_letter].width = 12
        ws.row_dimensions[HEADER_ROW].height = 32
        ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=3).coordinate

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

    nombre_cuartel = (cuartel.get("nombre") or cuartel.get("descripcion_ceco") or f"cuartel_{id_cuartel}")
    safe = "".join(c if c.isalnum() else "_" for c in str(nombre_cuartel))[:40]
    fn = f"programa_productos_{safe}.xlsx"
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


# ── Resumen de Costos ─────────────────────────────────────────────────────────

def _sucursales_permitidas_tuple(request: Request) -> tuple | None:
    """None => admin/super_admin (todas); tuple => solo esas."""
    perm = request.session.get("user_sucursales")
    if perm is None:
        return None
    return tuple(perm)


@app.get("/app/resumen-costos", response_class=HTMLResponse)
def web_resumen_costos(
    request: Request,
    temporada: str | None = None,
    sucursal: str | None = None,
    especie: str | None = None,
):
    temporada_id = _to_int(temporada)
    sucursal_id = _to_int(sucursal)
    especie_id = _to_int(especie)

    temporadas = get_temporadas()
    id_temp = temporada_id or (temporadas[0]["id"] if temporadas else None)
    sucursales_perm = _sucursales_permitidas_tuple(request)
    id_suc = _id_sucursal(request, sucursal_id)

    rows = get_resumen_costos(
        id_temporada=id_temp,
        sucursales_permitidas=sucursales_perm,
        id_sucursal=id_suc,
        id_especie=especie_id,
    )

    # Agrupar: especie -> variedad -> lista de cuarteles + subtotales
    from collections import defaultdict
    grupos: dict = {}
    for r in rows:
        esp = r["especie"]
        var = r["variedad"]
        grupos.setdefault(esp, {"variedades": {}, "sup": 0.0, "costo_ha_avg": 0.0, "costo_total": 0.0, "n": 0})
        grupos[esp]["variedades"].setdefault(var, {"id_variedad": r["id_variedad"], "cuarteles": [],
                                                    "sup": 0.0, "costo_total": 0.0, "n": 0})
        v_bucket = grupos[esp]["variedades"][var]
        cost_ha = float(r["costo_ha_total"] or 0)
        sup = float(r["sup_productiva"] or 0)
        cost_total = cost_ha * sup
        cuartel_dict = {
            "id_cuartel": r["id_cuartel"],
            "cuartel": r["cuartel"],
            "sucursal": r["sucursal"],
            "id_variedad": r["id_variedad"],
            "sup": sup,
            "costo_ha": cost_ha,
            "costo_total": cost_total,
            "n_semanas": r["n_semanas"],
            "tiene_dosis": (r["n_semanas_con_dosis"] or 0) > 0,
        }
        v_bucket["cuarteles"].append(cuartel_dict)
        v_bucket["sup"] += sup
        v_bucket["costo_total"] += cost_total
        v_bucket["n"] += 1
        grupos[esp]["sup"] += sup
        grupos[esp]["costo_total"] += cost_total
        grupos[esp]["n"] += 1

    # Calcular costo/ha promedio ponderado por sup
    for esp, g in grupos.items():
        g["costo_ha_avg"] = (g["costo_total"] / g["sup"]) if g["sup"] else 0
        for var, vg in g["variedades"].items():
            vg["costo_ha_avg"] = (vg["costo_total"] / vg["sup"]) if vg["sup"] else 0

    # Totales globales
    tot_sup = sum(g["sup"] for g in grupos.values())
    tot_cost = sum(g["costo_total"] for g in grupos.values())
    tot_n = sum(g["n"] for g in grupos.values())
    tot_costo_ha = (tot_cost / tot_sup) if tot_sup else 0

    # Catalogos filtros
    especies_all = get_especies(id_sucursal=id_suc)
    sucursales_all = get_sucursales()
    if sucursales_perm is not None:
        sucursales_all = [s for s in sucursales_all if s["id"] in sucursales_perm]

    return templates.TemplateResponse(
        "resumen_costos.html",
        {
            "request": request,
            "active_page": "resumen-costos",
            "temporadas": temporadas,
            "especies": especies_all,
            "sucursales": sucursales_all,
            "grupos": grupos,
            "filtro_temporada": id_temp,
            "filtro_sucursal": id_suc,
            "filtro_especie": especie_id,
            "tot_sup": tot_sup,
            "tot_cost": tot_cost,
            "tot_costo_ha": tot_costo_ha,
            "tot_n": tot_n,
        },
    )


@app.get("/app/resumen-costos/excel")
def resumen_costos_excel(
    request: Request,
    temporada: str | None = None,
    sucursal: str | None = None,
    especie: str | None = None,
):
    """Exporta el resumen de costos a Excel respetando los filtros aplicados."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt

    temporada_id = _to_int(temporada)
    sucursal_id = _to_int(sucursal)
    especie_id = _to_int(especie)

    temporadas = get_temporadas()
    id_temp = temporada_id or (temporadas[0]["id"] if temporadas else None)
    sucursales_perm = _sucursales_permitidas_tuple(request)
    id_suc = _id_sucursal(request, sucursal_id)

    rows = get_resumen_costos(
        id_temporada=id_temp,
        sucursales_permitidas=sucursales_perm,
        id_sucursal=id_suc,
        id_especie=especie_id,
    )

    # Etiquetas de los filtros aplicados
    temp_label = next((t["temporada"] for t in temporadas if t["id"] == id_temp), "")
    suc_label = "Todas"
    if id_suc:
        suc_row = next((s for s in get_sucursales() if s["id"] == id_suc), None)
        suc_label = suc_row["sucursal"] if suc_row else str(id_suc)
    esp_label = "Todas"
    if especie_id:
        esp_all = get_especies()
        esp_row = next((e for e in esp_all if e["id"] == especie_id), None)
        esp_label = esp_row["especie"] if esp_row else str(especie_id)

    # Estilos
    HDR_FILL = PatternFill("solid", fgColor="2d5a1f")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    ESP_FILL = PatternFill("solid", fgColor="A8C99A")
    VAR_FILL = PatternFill("solid", fgColor="D7E8C9")
    TOTAL_FILL = PatternFill("solid", fgColor="2d5a1f")
    TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
    ZEBRA = PatternFill("solid", fgColor="F4FAF1")
    THIN = Side(style="thin", color="C5D9BB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Costos"

    # Encabezado
    ws["A1"] = "Resumen de Costos por Cuartel - Fertilizaciones"
    ws["A1"].font = Font(bold=True, size=13, color="2d5a1f")
    ws.merge_cells("A1:G1")
    ws["A2"] = f"Temporada: {temp_label}  |  Sucursal: {suc_label}  |  Especie: {esp_label}  |  Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666", size=9)
    ws.merge_cells("A2:G2")

    # Headers tabla
    HDR_ROW = 4
    headers = ["Especie", "Variedad", "Cuartel", "Sucursal", "Sup. ha", "Costo/ha USD", "Total USD"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=HDR_ROW, column=i, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[HDR_ROW].height = 22

    # Agrupar
    from collections import defaultdict
    grupos: dict = {}
    for r in rows:
        esp = r["especie"]
        var = r["variedad"]
        grupos.setdefault(esp, {"variedades": {}, "sup": 0.0, "costo_total": 0.0, "n": 0})
        grupos[esp]["variedades"].setdefault(var, {"cuarteles": [], "sup": 0.0, "costo_total": 0.0})
        cost_ha = float(r["costo_ha_total"] or 0)
        sup = float(r["sup_productiva"] or 0)
        cost_total = cost_ha * sup
        grupos[esp]["variedades"][var]["cuarteles"].append({
            "cuartel": r["cuartel"], "sucursal": r["sucursal"],
            "sup": sup, "costo_ha": cost_ha, "costo_total": cost_total,
        })
        grupos[esp]["variedades"][var]["sup"] += sup
        grupos[esp]["variedades"][var]["costo_total"] += cost_total
        grupos[esp]["sup"] += sup
        grupos[esp]["costo_total"] += cost_total
        grupos[esp]["n"] += 1

    row = HDR_ROW + 1
    tot_sup = 0.0
    tot_cost = 0.0
    tot_n = 0

    for esp, esp_grp in grupos.items():
        zebra = False
        for var, var_grp in esp_grp["variedades"].items():
            for c in var_grp["cuarteles"]:
                vals = [esp, var, c["cuartel"], c["sucursal"],
                        round(c["sup"], 2),
                        round(c["costo_ha"], 2) if c["costo_ha"] > 0 else None,
                        round(c["costo_total"], 2) if c["costo_total"] > 0 else None]
                for i, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=i, value=v)
                    cell.border = BORDER
                    if i >= 5:
                        cell.alignment = Alignment(horizontal="right")
                    if zebra:
                        cell.fill = ZEBRA
                for col in (6, 7):
                    ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
                ws.cell(row=row, column=5).number_format = "0.00"
                row += 1
                zebra = not zebra

            # Subtotal por variedad
            var_costo_ha_avg = (var_grp["costo_total"] / var_grp["sup"]) if var_grp["sup"] else 0
            vals = ["", f"TOTAL {var}", "", "",
                    round(var_grp["sup"], 2),
                    round(var_costo_ha_avg, 2),
                    round(var_grp["costo_total"], 2)]
            for i, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=i, value=v)
                cell.fill = VAR_FILL
                cell.font = Font(bold=True)
                cell.border = BORDER
                if i >= 5:
                    cell.alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=5).number_format = "0.00"
            for col in (6, 7):
                ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
            row += 1

        # Subtotal por especie
        esp_costo_ha_avg = (esp_grp["costo_total"] / esp_grp["sup"]) if esp_grp["sup"] else 0
        vals = [f"TOTAL {esp}", "", "", "",
                round(esp_grp["sup"], 2),
                round(esp_costo_ha_avg, 2),
                round(esp_grp["costo_total"], 2)]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.fill = ESP_FILL
            cell.font = Font(bold=True, size=11)
            cell.border = BORDER
            if i >= 5:
                cell.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=5).number_format = "0.00"
        for col in (6, 7):
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
        row += 2  # espacio entre especies

        tot_sup += esp_grp["sup"]
        tot_cost += esp_grp["costo_total"]
        tot_n += esp_grp["n"]

    # Total global
    tot_costo_ha = (tot_cost / tot_sup) if tot_sup else 0
    vals = ["TOTAL GENERAL", f"{tot_n} cuarteles", "", "",
            round(tot_sup, 2), round(tot_costo_ha, 2), round(tot_cost, 2)]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER
        if i >= 5:
            cell.alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=5).number_format = "0.00"
    for col in (6, 7):
        ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

    # Anchos + freeze
    widths = [18, 18, 32, 20, 10, 13, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1).coordinate

    # Filename
    parts = ["ResumenCostos"]
    if suc_label != "Todas":
        parts.append("".join(ch if ch.isalnum() else "_" for ch in suc_label)[:20])
    if esp_label != "Todas":
        parts.append("".join(ch if ch.isalnum() else "_" for ch in esp_label)[:20])
    if temp_label:
        parts.append("".join(ch if ch.isalnum() else "_" for ch in temp_label))
    fn = "_".join(parts) + ".xlsx"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@app.get("/app/resumen-costos/comparar/{id_cuartel}", response_class=HTMLResponse)
def resumen_costos_comparar(
    request: Request,
    id_cuartel: int,
    id_variedad: int,
    temporada: str | None = None,
):
    """Fragmento HTMX: compara este cuartel contra los otros de misma variedad
    accesibles al usuario. Solo devuelve HTML del panel."""
    _require_cuartel_permitido(request, id_cuartel)
    id_temp = _to_int(temporada)
    sucursales_perm = _sucursales_permitidas_tuple(request)
    filas = get_comparativa_variedad(
        id_variedad=id_variedad,
        id_temporada=id_temp,
        sucursales_permitidas=sucursales_perm,
    )
    # Enriquecer con costo total y flag "actual"
    filas_e = []
    for f in filas:
        sup = float(f["sup_productiva"] or 0)
        cost_ha = float(f["costo_ha_total"] or 0)
        filas_e.append({
            "id_cuartel": f["id_cuartel"],
            "cuartel": f["cuartel"],
            "sucursal": f["sucursal"],
            "sup": sup,
            "costo_ha": cost_ha,
            "costo_total": cost_ha * sup,
            "es_actual": f["id_cuartel"] == id_cuartel,
        })
    return templates.TemplateResponse(
        "fragment_comparativa_variedad.html",
        {
            "request": request,
            "filas": filas_e,
            "id_cuartel_actual": id_cuartel,
            "filtro_temporada": id_temp,
        },
    )


# ── Parámetros ────────────────────────────────────────────────────────────────

@app.get("/app/parametros", response_class=HTMLResponse)
def web_parametros(request: Request):
    _require_admin(request)
    return templates.TemplateResponse(
        "parametros.html",
        {
            "request": request,
            "active_page": "parametros",
            "vigores": get_vigores(),
            "factores": get_factores_all(),
        },
    )


# ══ Gestion de usuarios (solo super_admin) ════════════════════════════════════

@app.get("/app/parametros/usuarios", response_class=HTMLResponse)
def web_usuarios(request: Request, ok: str | None = None, err: str | None = None):
    _require_super_admin(request)
    usuarios = listar_usuarios_con_sucursales()
    sucursales = request.state.sucursales_all  # admin/super_admin ven todas
    return templates.TemplateResponse(
        "usuarios.html",
        {
            "request": request,
            "active_page": "usuarios",
            "usuarios": usuarios,
            "sucursales": sucursales,
            "alert_ok": {"creado": "Usuario creado.",
                         "actualizado": "Usuario actualizado.",
                         "eliminado": "Usuario eliminado."}.get(ok),
            "alert_err": {"duplicado": "El nombre de usuario ya existe.",
                          "email_invalido": "El correo debe ser del dominio @lahornilla.cl.",
                          "self_delete": "No puedes eliminarte a ti mismo.",
                          "self_demote": "No puedes quitarte el rol super_admin a ti mismo."}.get(err),
        },
    )


@app.post("/app/parametros/usuarios")
def post_crear_usuario(
    request: Request,
    usuario: str = Form(...),
    nombre: str = Form(...),
    apellido: str = Form(...),
    email: str = Form(...),
    rol: str = Form("user"),
    id_sucursal: list[str] = Form(default=[]),
):
    _require_super_admin(request)
    usuario_n = (usuario or "").strip()
    email_n = (email or "").strip().lower()
    if not usuario_n or existe_usuario_nombre(usuario_n):
        return RedirectResponse(url="/app/parametros/usuarios?err=duplicado", status_code=303)
    if not email_n.endswith("@lahornilla.cl"):
        return RedirectResponse(url="/app/parametros/usuarios?err=email_invalido", status_code=303)
    new_id = crear_usuario(usuario_n, nombre, apellido, email_n, rol)
    if rol == "user" and id_sucursal:
        set_sucursales_usuario(new_id, id_sucursal)
    return RedirectResponse(url="/app/parametros/usuarios?ok=creado", status_code=303)


@app.post("/app/parametros/usuarios/{id_usuario}/rol")
def post_rol_usuario(
    request: Request,
    id_usuario: int,
    rol: str = Form(...),
):
    _require_super_admin(request)
    # No permitir quitarse el super_admin a uno mismo
    if id_usuario == int(request.session.get("user_id") or 0) and rol != "super_admin":
        return RedirectResponse(url="/app/parametros/usuarios?err=self_demote", status_code=303)
    try:
        actualizar_rol_usuario(id_usuario, rol)
    except ValueError:
        pass
    return RedirectResponse(url="/app/parametros/usuarios?ok=actualizado", status_code=303)


@app.post("/app/parametros/usuarios/{id_usuario}/sucursales")
def post_sucursales_usuario(
    request: Request,
    id_usuario: int,
    id_sucursal: list[str] = Form(default=[]),
):
    _require_super_admin(request)
    set_sucursales_usuario(id_usuario, id_sucursal)
    return RedirectResponse(url="/app/parametros/usuarios?ok=actualizado", status_code=303)


@app.post("/app/parametros/usuarios/{id_usuario}/email")
def post_email_usuario(
    request: Request,
    id_usuario: int,
    email: str = Form(...),
):
    _require_super_admin(request)
    email_n = (email or "").strip().lower()
    if not email_n.endswith("@lahornilla.cl"):
        return RedirectResponse(url="/app/parametros/usuarios?err=email_invalido", status_code=303)
    actualizar_email_usuario(id_usuario, email_n)
    return RedirectResponse(url="/app/parametros/usuarios?ok=actualizado", status_code=303)


@app.post("/app/parametros/usuarios/{id_usuario}/eliminar")
def post_eliminar_usuario(request: Request, id_usuario: int):
    _require_super_admin(request)
    if id_usuario == int(request.session.get("user_id") or 0):
        return RedirectResponse(url="/app/parametros/usuarios?err=self_delete", status_code=303)
    eliminar_usuario(id_usuario)
    return RedirectResponse(url="/app/parametros/usuarios?ok=eliminado", status_code=303)


@app.post("/app/parametros/vigor")
def post_vigor(
    request: Request,
    id: int | None = Form(None),
    vigor: str = Form(...),
    factor: float = Form(...),
):
    _require_admin(request)
    save_vigor(id, vigor, factor)
    return RedirectResponse(url="/app/parametros", status_code=303)


@app.post("/app/parametros/factor/{id_factor}")
def post_factor(
    request: Request,
    id_factor: int,
    factor_uva: float = Form(0),
    factor_cereza: float = Form(0),
    factor_ciruela: float = Form(0),
    factor_nectarin: float = Form(0),
    factor_durazno: float = Form(0),
    factor_damasco: float = Form(0),
):
    _require_admin(request)
    save_factor(
        id_factor,
        factor_uva=factor_uva,
        factor_cereza=factor_cereza,
        factor_ciruela=factor_ciruela,
        factor_nectarin=factor_nectarin,
        factor_durazno=factor_durazno,
        factor_damasco=factor_damasco,
    )
    return RedirectResponse(url="/app/parametros", status_code=303)


_TIPOS_HABILITADOS = [('5', 'Abono'), ('48', 'Fertilizante')]


@app.get("/app/parametros/productos", response_class=HTMLResponse)
def web_productos(request: Request, err: str | None = None, ok: str | None = None,
                  err_detail: str | None = None, tipo: str | None = None):
    _require_admin(request)
    tipo_activo = tipo if tipo in {t[0] for t in _TIPOS_HABILITADOS} else '5'
    con_nutrientes = (tipo_activo == '5')

    errores = {
        "duplicado": "Ya existe un producto con ese nombre. No se creó duplicado.",
        "en_uso": (
            f"No se puede eliminar: el producto está en uso en "
            f"{err_detail or '?'} dosis cargadas. Quitalo de la matriz primero."
        ),
    }
    oks = {
        "eliminado": "Producto eliminado correctamente.",
        "creado": "Producto creado correctamente.",
    }

    productos = get_productos_lista(id_actividad=tipo_activo)
    # Para productos no-fertilizantes, traer sus IAs cargados
    ias_por_producto: dict = {}
    if not con_nutrientes:
        for p in productos:
            ias_por_producto[p["id"]] = get_ias_de_producto(p["id"])

    return templates.TemplateResponse(
        "productos.html",
        {
            "request": request,
            "active_page": "productos",
            "tipos_habilitados": _TIPOS_HABILITADOS,
            "tipo_activo": tipo_activo,
            "con_nutrientes": con_nutrientes,
            "productos": productos,
            "ias_por_producto": ias_por_producto,
            "unidades": get_unidades_lista(),
            "objetivos": get_objetivos(),
            "modos_accion": get_modos_accion(),
            "ingredientes_activos": get_ingredientes_activos() if not con_nutrientes else [],
            "alert_err": errores.get(err) if err else None,
            "alert_ok": oks.get(ok) if ok else None,
        },
    )


@app.post("/app/parametros/productos")
def crear_producto(
    request: Request,
    nombre_comercial: str = Form(...),
    id_unidad: int = Form(...),
    codigo_softland: int | None = Form(None),
    precio_usd: float = Form(0),
    eficiencia: float = Form(100),
    id_objetivo: str = Form(""),
    id_modo_accion: str = Form(""),
    reingreso: str = Form(""),
    n: float = Form(0),
    k: float = Form(0),
    p: float = Form(0),
    mg: float = Form(0),
    b: float = Form(0),
    ca: float = Form(0),
    zn: float = Form(0),
    mn: float = Form(0),
    fe: float = Form(0),
    id_actividad: str = Form('5'),
):
    _require_admin(request)
    # Solo permitir tipos habilitados
    if id_actividad not in {t[0] for t in _TIPOS_HABILITADOS}:
        id_actividad = '5'
    # Anti-duplicado dentro del mismo tipo
    if existe_producto_por_nombre(nombre_comercial.strip(), id_actividad=id_actividad):
        return RedirectResponse(
            url=f"/app/parametros/productos?tipo={id_actividad}&err=duplicado",
            status_code=303,
        )
    # UI envia 0-100, BD guarda fracciones 0-1 (solo aplica a fertilizantes)
    reingreso_int = _to_int(reingreso)
    save_producto(
        nombre_comercial.strip(), id_unidad, codigo_softland,
        n/100, k/100, p/100, mg/100, b/100, ca/100, zn/100, mn/100,
        eficiencia/100,
        precio_usd=precio_usd,
        id_objetivo=id_objetivo or None,
        id_modo_accion=id_modo_accion or None,
        reingreso=reingreso_int,
        fe=fe/100,
        id_actividad=id_actividad,
    )
    return RedirectResponse(
        url=f"/app/parametros/productos?tipo={id_actividad}&ok=creado",
        status_code=303,
    )


@app.post("/app/parametros/productos/{id_producto}/general")
def editar_producto_general(
    request: Request,
    id_producto: str,
    nombre_comercial: str | None = Form(None),
    id_unidad: str | None = Form(None),
    codigo_softland: str | None = Form(None),
    precio_usd: float = Form(0),
    id_objetivo: str = Form(""),
    id_modo_accion: str = Form(""),
    reingreso: str = Form(""),
    id_actividad: str = Form('5'),
):
    """Edicion para productos NO fertilizantes (sin nutrientes)."""
    _require_admin(request)
    nombre = (nombre_comercial or "").strip()
    if nombre and existe_producto_por_nombre(nombre, excluir_id=id_producto, id_actividad=id_actividad):
        return RedirectResponse(
            url=f"/app/parametros/productos?tipo={id_actividad}&err=duplicado",
            status_code=303,
        )
    cod_set = codigo_softland is not None
    update_producto_general(
        id_producto,
        nombre_comercial=nombre if nombre else None,
        id_unidad=_to_int(id_unidad),
        codigo_softland=_to_int(codigo_softland) if cod_set else None,
        _codigo_softland_set=cod_set,
        precio_usd=precio_usd,
        id_objetivo=id_objetivo,
        id_modo_accion=id_modo_accion,
        reingreso=_to_int(reingreso),
    )
    return RedirectResponse(
        url=f"/app/parametros/productos?tipo={id_actividad}",
        status_code=303,
    )


@app.post("/app/parametros/productos/{id_producto}/ias")
def editar_producto_ias(
    request: Request,
    id_producto: str,
    id_actividad: str = Form('5'),
    id_ia: list[str] = Form(default=[]),
    porcentaje: list[str] = Form(default=[]),
    base_comparacion: list[str] = Form(default=[]),
):
    """Reemplaza la lista de IAs de un producto. Listas paralelas."""
    _require_admin(request)
    ias = []
    for i, iid in enumerate(id_ia):
        if not iid:
            continue
        try:
            pct = float(porcentaje[i]) if i < len(porcentaje) and porcentaje[i] else 0
        except (ValueError, TypeError):
            pct = 0
        base = base_comparacion[i] if i < len(base_comparacion) and base_comparacion[i] else "p/p"
        ias.append({"id_ia": iid, "porcentaje": pct, "base_comparacion": base})

    save_ias_de_producto(id_producto, ias)
    return RedirectResponse(
        url=f"/app/parametros/productos?tipo={id_actividad}",
        status_code=303,
    )


@app.post("/app/parametros/ias/nuevo")
def crear_ia_endpoint(request: Request, nombre: str = Form(...)):
    """Crea un nuevo ingrediente activo. Responde JSON para uso desde fetch()."""
    _require_admin(request)
    resultado = crear_ingrediente_activo(nombre)
    if "error" in resultado:
        return JSONResponse(resultado, status_code=400)
    return JSONResponse(resultado)


@app.post("/app/parametros/productos/{id_producto}/eliminar")
def eliminar_producto_endpoint(request: Request, id_producto: str):
    _require_admin(request)
    ok, en_uso = eliminar_producto_db(id_producto)
    if ok:
        return RedirectResponse(url="/app/parametros/productos?ok=eliminado", status_code=303)
    return RedirectResponse(
        url=f"/app/parametros/productos?err=en_uso&err_detail={en_uso}",
        status_code=303,
    )


@app.post("/app/parametros/productos/{id_producto}/nutrientes")
def editar_nutrientes(
    request: Request,
    id_producto: str,
    eficiencia: float = Form(100),
    precio_usd: float = Form(0),
    id_objetivo: str = Form(""),
    id_modo_accion: str = Form(""),
    reingreso: str = Form(""),
    nombre_comercial: str | None = Form(None),
    id_unidad: str | None = Form(None),
    codigo_softland: str | None = Form(None),
    n: float = Form(0),
    k: float = Form(0),
    p: float = Form(0),
    mg: float = Form(0),
    b: float = Form(0),
    ca: float = Form(0),
    zn: float = Form(0),
    mn: float = Form(0),
    fe: float = Form(0),
):
    _require_admin(request)
    # UI envia 0-100, BD guarda fracciones 0-1
    reingreso_int = _to_int(reingreso)

    # Validar duplicado de nombre si esta cambiando
    nombre = (nombre_comercial or "").strip()
    if nombre and existe_producto_por_nombre(nombre, excluir_id=id_producto):
        return RedirectResponse(
            url="/app/parametros/productos?err=duplicado",
            status_code=303,
        )

    id_unidad_int = _to_int(id_unidad)
    cod_set = codigo_softland is not None
    cod_int = _to_int(codigo_softland) if cod_set else None

    update_producto_nutrientes(
        id_producto,
        n/100, k/100, p/100, mg/100, b/100, ca/100, zn/100, mn/100,
        eficiencia/100,
        precio_usd=precio_usd,
        id_objetivo=id_objetivo,
        id_modo_accion=id_modo_accion,
        reingreso=reingreso_int,
        fe=fe/100,
        nombre_comercial=nombre if nombre else None,
        id_unidad=id_unidad_int,
        codigo_softland=cod_int,
        _codigo_softland_set=cod_set,
    )
    return RedirectResponse(url="/app/parametros/productos", status_code=303)


# ══ API DE PAPELETAS ══════════════════════════════════════════════════════════

@app.get("/papeleta/{id_programa}")
def generar_papeleta(request: Request, id_programa: str):
    programa = get_programa(id_programa)
    if not programa:
        raise HTTPException(status_code=404, detail=f"Programa '{id_programa}' no encontrado")

    # Bloquear si la sucursal del cuartel no esta entre las permitidas del usuario
    permitidas = request.session.get("user_sucursales")
    if permitidas is not None and programa.get("id_sucursal") not in set(permitidas):
        raise HTTPException(status_code=403, detail="Sucursal no autorizada")

    productos = get_productos(id_programa)
    if not productos:
        raise HTTPException(status_code=404, detail="El programa no tiene productos registrados")

    sectores = get_sectores(programa["id_cuartel"])
    pdf_bytes = build_pdf(programa, productos, sectores)

    filename = f"papeleta_{programa['etiqueta_semana']}_{id_programa[:8]}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/app/descuento-bodega", response_class=HTMLResponse)
def web_descuento_bodega(
    request: Request,
    temporada: str | None = None,
    sucursal: str | None = None,
    semana: str | None = None,
):
    temporada_id = _to_int(temporada)
    sucursal_id = _to_int(sucursal)

    temporadas = get_temporadas()
    id_temp = temporada_id or (temporadas[0]["id"] if temporadas else None)
    id_suc = _id_sucursal(request, sucursal_id)

    semanas = get_semanas_disponibles(id_temporada=id_temp, id_sucursal=id_suc)
    etiqueta_semana = semana or (semanas[0]["etiqueta_semana"] if semanas else None)

    filas = []
    resumen = []
    if etiqueta_semana:
        filas = get_descuento_bodega_semana(
            etiqueta_semana=etiqueta_semana,
            id_sucursal=id_suc,
            id_temporada=id_temp,
        )
        agg: dict = {}
        for f in filas:
            k = (f["id_producto"], f["producto"], f.get("codigo_softland"), f.get("unidad") or "")
            agg[k] = agg.get(k, 0.0) + float(f["cantidad_total"] or 0)
        resumen = [
            {
                "id_producto": k[0],
                "producto": k[1],
                "codigo_softland": k[2],
                "unidad": k[3],
                "cantidad_total": round(v, 2),
            }
            for k, v in sorted(agg.items(), key=lambda x: x[0][1])
        ]

    return templates.TemplateResponse(
        "descuento_bodega.html",
        {
            "request": request,
            "active_page": "descuento-bodega",
            "temporadas": temporadas,
            "semanas": semanas,
            "filas": filas,
            "resumen": resumen,
            "filtro_temporada": id_temp,
            "filtro_sucursal": id_suc,
            "filtro_semana": etiqueta_semana,
        },
    )


@app.get("/app/descuento-bodega/excel")
def descargar_descuento_bodega_excel(
    request: Request,
    temporada: str | None = None,
    sucursal: str | None = None,
    semana: str = "",
):
    if not semana:
        raise HTTPException(status_code=400, detail="Falta el parametro 'semana'")
    temporada_id = _to_int(temporada)
    sucursal_id = _to_int(sucursal)
    id_suc = _id_sucursal(request, sucursal_id)

    filas = get_descuento_bodega_semana(
        etiqueta_semana=semana,
        id_sucursal=id_suc,
        id_temporada=temporada_id,
    )

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HDR_FILL = PatternFill("solid", fgColor="2d5a1f")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    ZEBRA = PatternFill("solid", fgColor="F4FAF1")
    THIN = Side(style="thin", color="C5D9BB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Detalle"

    headers = [
        "Semana", "Fecha inicio", "Fecha fin",
        "Sucursal", "Cuartel (CECO id)", "Cuartel", "Variedad",
        "Sup. ha", "Producto", "Cód. Softland", "Unidad",
        "Dosis/ha", "Cantidad total",
    ]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    for i, f in enumerate(filas, start=2):
        ws1.append([
            f["etiqueta_semana"],
            f["fecha_inicio"],
            f["fecha_fin"],
            f["sucursal"],
            f["id_cuartel"],
            f["cuartel"],
            f["variedad"],
            float(f["sup_productiva"] or 0),
            f["producto"],
            f["codigo_softland"] or "",
            f["unidad"],
            float(f["dosis_ha"] or 0),
            float(f["cantidad_total"] or 0),
        ])
        if i % 2 == 0:
            for cell in ws1[i]:
                cell.fill = ZEBRA
        for cell in ws1[i]:
            cell.border = BORDER

    widths = [16, 12, 12, 22, 18, 32, 22, 9, 32, 16, 8, 13, 14]
    for col_idx, w in enumerate(widths, start=1):
        ws1.column_dimensions[ws1.cell(row=1, column=col_idx).column_letter].width = w
    ws1.freeze_panes = "A2"

    # Hoja 2: resumen consolidado
    ws2 = wb.create_sheet("Resumen bodega")
    agg: dict = {}
    for f in filas:
        k = (f["producto"], f.get("codigo_softland"), f.get("unidad") or "")
        agg[k] = agg.get(k, 0.0) + float(f["cantidad_total"] or 0)

    ws2.append(["Producto", "Cód. Softland", "Unidad", "Cantidad total"])
    for cell in ws2[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    for i, (k, v) in enumerate(sorted(agg.items()), start=2):
        ws2.append([k[0], k[1] or "", k[2], round(v, 2)])
        if i % 2 == 0:
            for cell in ws2[i]:
                cell.fill = ZEBRA
        for cell in ws2[i]:
            cell.border = BORDER

    widths2 = [32, 16, 10, 14]
    for col_idx, w in enumerate(widths2, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = w
    ws2.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_sem = semana.replace(" ", "_").replace("/", "-")
    filename = f"descuento_bodega_{safe_sem}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/app/adquisiciones", response_class=HTMLResponse)
def web_adquisiciones(
    request: Request,
    temporada: str | None = None,
    sucursal: str | None = None,
    semana_desde: str | None = None,
    semana_hasta: str | None = None,
):
    temporada_id = _to_int(temporada)
    sucursal_id = _to_int(sucursal)
    semana_desde_id = _to_int(semana_desde)
    semana_hasta_id = _to_int(semana_hasta)

    temporadas = get_temporadas()
    id_temp = temporada_id or (temporadas[0]["id"] if temporadas else None)
    id_suc = _id_sucursal(request, sucursal_id)

    semanas = get_semanas_temporada(id_temp) if id_temp else []

    filas = []
    total_kg = 0.0
    total_usd = 0.0
    totales_por_unidad: dict = {}
    if id_temp:
        filas = get_adquisiciones_consolidado(
            id_temporada=id_temp,
            id_semana_desde=semana_desde_id,
            id_semana_hasta=semana_hasta_id,
            id_sucursal=id_suc,
        )
        for f in filas:
            kg = float(f["cantidad_total"] or 0)
            f["subtotal_usd"] = round(kg * float(f.get("precio_usd") or 0), 2)
            total_kg += kg
            total_usd += f["subtotal_usd"]
            uni = f.get("unidad") or "kg"
            totales_por_unidad[uni] = totales_por_unidad.get(uni, 0) + kg

    return templates.TemplateResponse(
        "adquisiciones.html",
        {
            "request": request,
            "active_page": "adquisiciones",
            "temporadas": temporadas,
            "semanas": semanas,
            "filas": filas,
            "total_kg": round(total_kg, 2),
            "total_usd": round(total_usd, 2),
            "totales_por_unidad": {u: round(v, 2) for u, v in totales_por_unidad.items()},
            "filtro_temporada": id_temp,
            "filtro_sucursal": id_suc,
            "filtro_semana_desde": semana_desde_id,
            "filtro_semana_hasta": semana_hasta_id,
        },
    )


@app.get("/app/adquisiciones/excel")
def descargar_adquisiciones_excel(
    request: Request,
    temporada: str | None = None,
    sucursal: str | None = None,
    semana_desde: str | None = None,
    semana_hasta: str | None = None,
):
    temporada_id = _to_int(temporada)
    sucursal_id = _to_int(sucursal)
    semana_desde_id = _to_int(semana_desde)
    semana_hasta_id = _to_int(semana_hasta)

    temporadas = get_temporadas()
    id_temp = temporada_id or (temporadas[0]["id"] if temporadas else None)
    id_suc = _id_sucursal(request, sucursal_id)
    if not id_temp:
        raise HTTPException(status_code=400, detail="Falta temporada")

    filas = get_adquisiciones_consolidado(
        id_temporada=id_temp,
        id_semana_desde=semana_desde_id,
        id_semana_hasta=semana_hasta_id,
        id_sucursal=id_suc,
    )

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HDR_FILL = PatternFill("solid", fgColor="2d5a1f")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    TOTAL_FILL = PatternFill("solid", fgColor="b8dca4")
    TOTAL_FONT = Font(bold=True, color="1a3a10", size=11)
    ZEBRA = PatternFill("solid", fgColor="F4FAF1")
    THIN = Side(style="thin", color="C5D9BB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = Workbook()
    ws = wb.active
    ws.title = "Adquisiciones"

    # Cabecera con metadata del filtro
    semanas = get_semanas_temporada(id_temp)
    sem_desde = next((s for s in semanas if s["id"] == semana_desde_id), None)
    sem_hasta = next((s for s in semanas if s["id"] == semana_hasta_id), None)
    suc_info = get_sucursal_info(id_suc) if id_suc else None
    temp_info = next((t for t in temporadas if t["id"] == id_temp), None)

    ws.append(["Resumen de adquisiciones"])
    ws["A1"].font = Font(bold=True, color="2d5a1f", size=14)
    ws.append([])
    ws.append(["Temporada:", temp_info["temporada"] if temp_info else "—"])
    ws.append(["Sucursal:", suc_info["sucursal"] if suc_info else "Todas"])
    ws.append([
        "Rango:",
        f"{sem_desde['etiqueta_semana'] if sem_desde else 'Desde inicio'} → {sem_hasta['etiqueta_semana'] if sem_hasta else 'Hasta fin'}",
    ])
    ws.append([])

    headers = ["Producto", "Cód. Softland", "Unidad", "Cantidad total", "Precio USD/u", "Subtotal USD"]
    ws.append(headers)
    row_header = ws.max_row
    for cell in ws[row_header]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    total_kg = 0.0
    total_usd = 0.0
    for f in filas:
        kg = float(f["cantidad_total"] or 0)
        precio = float(f.get("precio_usd") or 0)
        subtotal = round(kg * precio, 2)
        ws.append([
            f["producto"],
            f.get("codigo_softland") or "",
            f.get("unidad") or "",
            kg,
            precio,
            subtotal,
        ])
        total_kg += kg
        total_usd += subtotal
        i = ws.max_row
        if (i - row_header) % 2 == 0:
            for cell in ws[i]:
                cell.fill = ZEBRA
        for cell in ws[i]:
            cell.border = BORDER

    # Fila TOTAL
    ws.append(["TOTAL", "", "", round(total_kg, 2), "", round(total_usd, 2)])
    i = ws.max_row
    for cell in ws[i]:
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER

    widths = [34, 16, 10, 16, 14, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=row_header, column=col_idx).column_letter].width = w
    ws.freeze_panes = f"A{row_header + 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    rango = ""
    if sem_desde and sem_hasta:
        rango = f"_{sem_desde['etiqueta_semana']}_a_{sem_hasta['etiqueta_semana']}".replace(" ", "_")
    suc_txt = f"_{suc_info['sucursal'].replace(' ', '_')}" if suc_info else ""
    filename = f"adquisiciones{suc_txt}{rango}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/papeleta-casetas/{etiqueta_semana}/{id_sucursal}")
def listar_casetas_papeleta(request: Request, etiqueta_semana: str, id_sucursal: int):
    """Devuelve JSON con casetas que tienen programa esta semana. Para popular
    el select 'Papeleta por caseta' en programas.html."""
    permitidas = request.session.get("user_sucursales")
    if permitidas is not None and id_sucursal not in set(permitidas):
        raise HTTPException(status_code=403, detail="Sucursal no autorizada")
    casetas = get_casetas_con_programa(etiqueta_semana, id_sucursal)
    return JSONResponse([{"id": c["id"], "caseta": c["caseta"]} for c in casetas])


@app.get("/papeleta-caseta/{etiqueta_semana}/{id_sucursal}/{id_caseta}")
def generar_papeleta_caseta(
    request: Request, etiqueta_semana: str, id_sucursal: int, id_caseta: int,
):
    permitidas = request.session.get("user_sucursales")
    if permitidas is not None and id_sucursal not in set(permitidas):
        raise HTTPException(status_code=403, detail="Sucursal no autorizada")
    sucursal = get_sucursal_info(id_sucursal)
    if not sucursal:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")
    caseta = get_caseta_info(id_caseta)
    if not caseta or caseta["id_sucursal"] != id_sucursal:
        raise HTTPException(status_code=404, detail="Caseta no encontrada en esta sucursal")
    semana = get_semana_info(etiqueta_semana)
    if not semana:
        raise HTTPException(status_code=404, detail=f"Semana '{etiqueta_semana}' no encontrada")

    rows = get_papeleta_campo_rows(etiqueta_semana, id_sucursal)
    supervisor = request.session.get("user_name", "") if hasattr(request, "session") else ""
    pdf_bytes = build_pdf_campo(
        rows=rows,
        orfanos=[],
        sucursal=sucursal,
        semana=semana,
        supervisor=supervisor,
        id_caseta_filtro=id_caseta,
        caseta_nombre=caseta["caseta"],
    )
    safe_suc = sucursal["sucursal"].replace(" ", "_")
    safe_cas = caseta["caseta"].replace(" ", "_")
    safe_sem = etiqueta_semana.replace(" ", "_")
    filename = f"papeleta_{safe_suc}_{safe_cas}_{safe_sem}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@app.get("/papeleta-campo/{etiqueta_semana}/{id_sucursal}")
def generar_papeleta_campo(request: Request, etiqueta_semana: str, id_sucursal: int):
    # Bloquear sucursales fuera del set de permisos del usuario
    permitidas = request.session.get("user_sucursales")
    if permitidas is not None and id_sucursal not in set(permitidas):
        raise HTTPException(status_code=403, detail="Sucursal no autorizada")
    sucursal = get_sucursal_info(id_sucursal)
    if not sucursal:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")

    semana = get_semana_info(etiqueta_semana)
    if not semana:
        raise HTTPException(status_code=404, detail=f"Semana '{etiqueta_semana}' no encontrada")

    rows = get_papeleta_campo_rows(etiqueta_semana, id_sucursal)
    orfanos = get_cuarteles_huerfanos(etiqueta_semana, id_sucursal)

    supervisor = request.session.get("user_name", "") if hasattr(request, "session") else ""

    pdf_bytes = build_pdf_campo(
        rows=rows,
        orfanos=orfanos,
        sucursal=sucursal,
        semana=semana,
        supervisor=supervisor,
    )

    safe_suc = sucursal["sucursal"].replace(" ", "_")
    safe_sem = etiqueta_semana.replace(" ", "_")
    filename = f"papeleta_{safe_suc}_{safe_sem}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@app.get("/registro-semanal/{etiqueta_semana}")
def generar_papeleta_bodega(etiqueta_semana: str, pro: bool = False):
    programas = get_programas_semana(etiqueta_semana)
    if not programas:
        raise HTTPException(status_code=404, detail=f"No hay programas para la semana '{etiqueta_semana}'")

    ids_programa = [p["id"] for p in programas]
    ids_cuartel  = list({p["id_cuartel"] for p in programas})

    productos_rows = get_productos_multiples(ids_programa)
    sectores_rows  = get_sectores_multiples(ids_cuartel)

    productos_map: dict = {}
    for row in productos_rows:
        productos_map.setdefault(row["id_programa"], []).append(row)

    sectores_map: dict = {}
    for row in sectores_rows:
        sectores_map.setdefault(row["id_cuartel"], []).append(row)

    pdf_bytes = build_pdf_bodega(etiqueta_semana, programas, productos_map, sectores_map, pro=pro)

    suffix   = "_pro" if pro else ""
    filename = f"bodega_{etiqueta_semana}{suffix}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
